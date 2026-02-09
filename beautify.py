import openpyxl
from openpyxl.styles import Font, PatternFill
import pandas as pd

csv_file = pd.read_csv("2026-01-30T09_07_28.865Z_Q3j.csv", encoding="utf-8")
csv_file.to_excel("sati.xlsx", index=False)

file = openpyxl.load_workbook("sati.xlsx")
sheet = file.active

row_num = sheet.max_row
col_num = sheet.max_column

# imena_kolona = [sheet.cell(row=1, column=i).value for i in range(1, col_num+1)]


def delete_columns(col_list):
    imena_kolona = {cell.value: cell.column for cell in sheet[1]}
    indexi_kolona_za_brisanje = [imena_kolona.get(x) for x in col_list]
    indexi_kolona_za_brisanje.reverse()

    for i in indexi_kolona_za_brisanje:
        sheet.delete_cols(i, 1)


def format_col_names(col_list):
    # treba pronaci sve kolone u kojima ima text i onda njih promijenit umjesto da se hardkodira lista
    for n in col_list:
        sheet[n].font = Font(color="FFFFFF", bold=True, size=12)
        sheet[n].fill = PatternFill("solid", start_color="215C98")


def extract_date(datum):
    novi_dict_kolona = {cell.value: cell.column for cell in sheet[1]}
    index_kolone = novi_dict_kolona.get(f"{datum}")

    for r in range(2, row_num):
        sheet.cell(row=r, column=index_kolone).value = sheet.cell(
            row=r, column=index_kolone
        ).value.split(",")[0]


def set_col_width():
    indexi_kolona = [cell.column for cell in sheet[1]]
    for c in indexi_kolona:
        slovo_kolona = openpyxl.utils.cell.get_column_letter(c)
        length = 0
        for r in range(1, row_num):
            length = max(length, len(str(sheet.cell(row=r, column=c).value)))
        sheet.column_dimensions[slovo_kolona].width = length + 2


def time_tracked():
    novi_dict_kolona = {cell.value: cell.column for cell in sheet[1]}
    index_kolone = novi_dict_kolona.get("Time Tracked")

    for r in range(2, row_num):
        sheet.cell(row=r, column=index_kolone).value = (
            int(sheet.cell(row=r, column=index_kolone).value) / 3600000
        )


kolone_za_brisanje = [
    "User ID",
    "Time Entry ID",
    "Start",
    "Stop",
    "Stop Text",
    "Space ID",
    "Folder ID",
    "List ID",
    "List Name",
    "Task ID",
    "Task Status",
    "Due Date",
    "Due Date Text",
    "Start Date",
    "Start Date Text",
    "Task Time Estimated",
    "Task Time Estimated Text",
    "Task Time Spent",
    "Task Time Spent Text",
    "User Total Time Estimated",
    "User Total Time Estimated Text",
    "User Total Time Tracked",
    "User Total Time Tracked Text",
    "Tags",
    "Checklists",
    "User Period Time Spent",
    "User Period Time Spent Text",
    "Date Created",
    "Date Created Text",
    "Custom Task ID",
    "Parent Task ID",
]

col_names = [
    "A1",
    "B1",
    "C1",
    "D1",
    "E1",
    "F1",
    "G1",
    "H1",
    "I1",
    "J1",
    "K1",
    "L1",
    "M1",
    "N1",
]


delete_columns(kolone_za_brisanje)
set_col_width()
time_tracked()
extract_date("Start Text")
format_col_names(col_names)

file.save("sati.xlsx")
